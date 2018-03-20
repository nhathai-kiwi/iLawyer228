#!/usr/bin/env python
# -*- coding: utf-8 -*-
# iLawyer basic functions

import numpy as np
import openpyxl
import iLawyer_scikit as isk
from collections import OrderedDict
import os
import sys
reload(sys)
sys.setdefaultencoding('utf8')
dict
# file dictionary for check question has at least 2 key words
dict001 = 'dict001.txt'

# delete file has a name inp_fileName
def process_delete_file(inp_file_name):
    cmd = "rm -rf " + inp_file_name
    os.system(cmd)
    return 0
# DONE


# read data in id_column from inp_xlsx from the second row
def gen_column_from_xlsx(inp_xlsx, num_rows, id_column):
    data = []
    book = openpyxl.load_workbook(inp_xlsx)
    sheet = book.active
    for row in range(2, num_rows + 1): # ignore 1st row which is preserved for column title
        datum = sheet.cell(row=row, column=id_column).value
        data.append(datum)
    return data
# DONE


# print content of a prediction to txt file
# by default articles are stored in 2nd colums from the second row
# get article has id = answer_number --> return index( row = answer_number + 1, colums = id_column)
# TODO explain default value inp_xlsx = 'lawInEnterprise.xlsx'
def print_txt_from_prediction(inp_xlsx, answer_number, id_column, out_txt):
    # get article: answer_number from inp_xlsx file
    book = openpyxl.load_workbook(inp_xlsx)
    sheet = book.active
    article = sheet.cell(row = answer_number + 1, column = id_column).value

    # write article to txt file
    f = open(out_txt, "w")
    f.write(article)
    f.close()
    return 0
# DONE

# gen article has id = answer_number
def gen_article_from_prediction(inp_xlsx, answer_number, id_column):
    book = openpyxl.load_workbook(inp_xlsx)
    sheet = book.active
    article = sheet.cell(row=answer_number + 1, column=id_column).value
    return article
# DONE

# print array to txt file, each element in a line
def print_txt_from_array(string_array, out_txt):
    f = open(out_txt, "w")
    for each_string in string_array:
        f.write(str(each_string) + "\n")
    f.close()
    return 0
# DONE


# process vncore
# input: file name of a txt file
# output: TODO add description here
def process_vncore(inp_txt, out_txt):
    cmd = "java -Xmx2g -jar VnCoreNLP-1.0.jar -fin " + inp_txt + " -fout " + out_txt
    os.system(cmd)
    return 0
# DONE


# generate string array from txt file
def gen_array_from_txt(inp_txt):
    string_array = []
    f = open(inp_txt, 'r')
    for each_string in f:
        string_array.append(each_string)
    f.close()
    return string_array
# DONE


# remove identical element from an array
def gen_distinct_array(inp_arr):
    out_arr = []
    dicts = {}
    for element in inp_arr:
        value = dicts.get(element, 0)
        dicts[element] = value + 1
    for key, value in dicts.iteritems():
        out_arr.append(key)
    return out_arr
# DONE


# remove identical element from an array
def gen_distinct_array_with_count(inp_arr):
    out_arr = []
    dicts = {}
    for element in inp_arr:
        value = dicts.get(element, 0)
        dicts[element] = value + 1

    # print "Len dicts: ", len(dicts)
    dictsSort = OrderedDict(sorted(dicts.items(), key=lambda x: x[1], reverse=True))
    # print "Len dictSort: ", len(dictsSort)

    for key, value in dictsSort.items():
        out_arr.append(key + " " + str(value))
        #print "Here: ", key, " ", value
    return out_arr
# DONE


# generate words (string array) from output of vncore
# manipulation includes: ignore all words starting with digit,
# replace underscore by space and change all characters to lower case
def gen_words_from_vncore_out_txt(vn_core_out_txt):
    words_replicated = []
    lines = gen_array_from_txt(vn_core_out_txt)
    for line in lines:
        texts = line.split(chr(9))
        if len(texts) == 6:
            text = texts[1]
            if not text[0].isdigit():  # ignore all words starting with digit
                word = text.replace('_', ' ').lower()  # replace underscore by space, change all characters to lower case
                words_replicated.append(word)
    # replicated words will be handled in gen_distinct_array
    return words_replicated


# generate words (string array) from output of vncore
# manipulation includes those listed in gen_words_from_vncore_out_txt
# keep only words with multiple syllables
def gen_words_from_vncore_out_txt_02(vn_core_out_txt):
    words_replicated = []
    lines = gen_array_from_txt(vn_core_out_txt)
    for line in lines:
        texts = line.split(chr(9))
        if len(texts) == 6:
            text = texts[1]
            if not text[0].isdigit(): # ignore all words starting with digit
                if text.find('_') != -1:
                    word = text.replace('_', ' ').lower() # replace underscore by space, change all characters to lower case
                    words_replicated.append(word)
    return words_replicated


# generate feature vector (numeric array)
def gen_feature_vector(string_array):
    global dict
    feature_vector = []
    dict_string_array = {} # remove all the string duplicate in string_array by dictinary - STL python
    for element in string_array:
        value = dict_string_array.get(element, 0)
        dict_string_array[element] = value + 1
    for word in dict:
        if word in dict_string_array:
            feature_vector.append(1)
        else:
            feature_vector.append(0)
    return feature_vector
# DONE


# generate feature vector (numeric 1D array) from a question
def gen_input_vector_from_txt(inp_array):

    print_txt_from_array(inp_array, 'vncore_inp.txt')
    process_vncore('vncore_inp.txt', 'vncore_out.txt')
    string_array = gen_words_from_vncore_out_txt('vncore_out.txt')

    inp_vector = gen_feature_vector(string_array)

    process_delete_file('vncore_inp.txt')
    process_delete_file('vncore_out.txt')

    return inp_vector
# DONE


# generate word from inp_txt, return list string
def gen_string_array_from_txt(inp_txt):

    process_vncore(inp_txt, 'vncore_out.txt')
    string_array = gen_words_from_vncore_out_txt('vncore_out.txt')

    return string_array

# DONE


# check question have at least 2 keywords (keyword from file dict001.txt contains 241 words)
def check_question_at_least_2_keywords(string_array):
    words = gen_array_from_txt(dict001)

    # keyword read from file then the last character's keyword is '\n'
    # remove '\n' character
    keywords = []
    for word in words:
        keywords.append(word[:-1])

    # counting number of string_array list in keywords
    cnt = 0
    for word in string_array:
        if word in keywords:
            print "word in keyword: ", word
            cnt += 1
    return cnt > 1


# generate feature table (numeric 2D array) from a training set stored in xlsx
# TODO rename function
def gen_feature_table_from_xlsx(inp_xlsx, num_rows, id_column_question, dictMain):
    separator = 'separator'
    feature_table = []
    all_questions = gen_column_from_xlsx(inp_xlsx, num_rows, id_column_question)
    global dict
    dict = dictMain

    questions_with_seperator = []
    for question in all_questions:
        # add 'separator' before each question for convenient process
        questions_with_seperator.append(separator)
        questions_with_seperator.append(question)

    print_txt_from_array(questions_with_seperator, 'vncore_inp.txt')
    process_vncore('vncore_inp.txt', 'vncore_out.txt')
    string_array = gen_words_from_vncore_out_txt('vncore_out.txt')

    # separating question and get its feature vector
    question = []
    cnt = 0
    for each_string in string_array:
        if each_string == separator:
            if len(question) == 0:
                continue
            feature_vector = gen_feature_vector(question)
            feature_table.append(feature_vector)
            question = []
        else:
            question.append(each_string)

    # add a last feature_vector question
    feature_vector = gen_feature_vector(question)
    feature_table.append(feature_vector)

    process_delete_file('vncore_inp.txt')
    process_delete_file('vncore_out.txt')

    return feature_table
# DONE


# test some of pair (alpha, number of iteration) im MLP model to choose the best (alpha, iteration)
def test_alpha_and_iteration_in_MLP(X, y, X1, y1, alphas, max_iters, num_features):
    for alpha in alphas:
        for max_iter in max_iters:
            print "Alpha: ", alpha, "max_iter: ", max_iter
            clf = isk.train_by_MLPClassifier_regularization(X, y, num_features, alpha, max_iter)

            cal_labels = isk.predict(clf, X)
            correct_labels = y
            performance = isk.cal_performance(correct_labels, cal_labels)
            print "Performances MLP training set: ", performance

            cal_labels = isk.predict(clf, X1)
            correct_labels = y1

            performance = isk.cal_performance(correct_labels, cal_labels)
            print "Performances MLP test set: ", performance
# DONE

