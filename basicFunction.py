#!/usr/bin/env python
# -*- coding: utf-8 -*- 

import os
import sys  
reload(sys)  
sys.setdefaultencoding('utf8')
import openpyxl

# def writeTextIntoFile():
# 	return 'writeTextIntoFile.txt'

fileNameForRunVnCore ='fileNameForRunVnCore.txt'
fileNameForRunVnCorewordInFormwordList = 'fileNameForRunVnCorewordInFormwordList.txt'

#return lines contains array of line in fileName
def getTextFromFile(fileName):
	lines = []
	f = open(fileName, 'r')
	for line in f:
		lines.append(line)
	f.close()
	return lines

#return number of distinct words from fileInput
def countingDistinctWords(fileInput):
	lines = getTextFromFile(fileInput)
	dicts  = {}
	for line in lines:
		#print line
		value = dicts.get(line, 0)
		dicts[line] = value + 1
	print "Number words distinict: ", len(dicts)

#return distinict element from array //remove all element begin with first character is digit
def distinctWords(Array):
	dicts = {}
	for element in Array:
		if (element[0].isdigit() == False):
			value = dicts.get(element, 0)
			dicts[element] = value + 1
	return dicts

#read xlsx file fileName, row: numberOfRows, column: numberofCols

def readFileXlsx(fileName, numberOfRows, numberOfCols):
	print "ReadfileXlsx from basicFunction"
	book = openpyxl.load_workbook(fileName)
	sheet = book.active
	ret = []
	#f = open('vnTokenizer.txt', "w")
	for row in range(1, numberOfRows  + 1):
		rowValue = []
		for col in range(1, numberOfCols + 1):	
			index = sheet.cell(row = row, column = col).value
			rowValue.append(index)
			# mainText = sheet.cell(row = row, column = 2).value
			# ret = processText(ith, mainText)
			# f.write(ret)
		ret.append(rowValue)
	return ret

#write data from .xlsx file to .txt file: each element in table written in one line in .txt file
def writeDataXlxsToTxt(fileXlsx, rows, cols):
	fileTxt = fileXlsx.replace('.xlsx', '.txt')
	f = open(fileTxt, "w")
	ret = readFileXlsx(fileXlsx, rows, cols)
	for row in ret:
		for col in row:
			f.write(str(col) + "\n")
	f.close()

#write array into fileName file
def writeArrayToTxtFile(fileName, arr):
	f = open(fileName, "w")
	for i in arr:
		f.write(str(i) + "\n")
	f.close()

#return vector keyword from file keyword.txt
def getkeyWord():
	return getTextFromFile('keyword.txt')

#return article in law of enterprise which correspond to id
def getArticle(fileName, numberOfRows, numberOfColums, id):
	tableLaw = readFileXlsx(fileName, numberOfRows, numberOfColums)
	return "Điều " + str(id) + " Luật Doanh Nghiệp 2014:\n" + tableLaw[id][1]

#write string value into fileName
def printOutput(fileName, value):
	f = open(fileName, 'w')
	f.write(value)
	f.close()

if __name__ == '__main__':
	print "basicFunction"
	article = getArticle('law.xlsx', 214, 2, 5)
	print article
	printOutput('answer.txt', article)